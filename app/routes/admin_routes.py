from collections import defaultdict
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from datetime import datetime
from flask_login import login_required, current_user
from sqlalchemy import extract
from app import db
from datetime import datetime, date
import calendar
from app.models import Penalty, User, Company, Attendance, Increment, db, Clearance, Broadcast
from werkzeug.security import generate_password_hash
from sqlalchemy.orm import joinedload



admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ============================================================
# 🔹 Admin Dashboard
# ============================================================
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    
    if getattr(current_user, "role", None) != "admin":
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for('auth.login'))
    today = date.today()
    total_agents = User.query.filter(User.role.ilike('agent')).count()
    disabled_users = User.query.filter_by(is_active_db=False).count()
    total_bonuses = 0
    company_count = Company.query.count()
    
    stats = {
        "total_agents": total_agents,
        "disabled_users": disabled_users,
        "total_bonuses": total_bonuses,
        "company_count": company_count
    }
    current_date=today
    return render_template(
        'admin/dashboard.html',
        stats=stats,
        current_date=today
    )

# ============================================================
# 🔹 Manage Users
# ============================================================
@admin_bp.route('/manage-users')
@login_required
def manage_users():
    if getattr(current_user, "role", None) != "admin":
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for('auth.login'))

    # eager-load Company to avoid N+1 queries when rendering company.name
    users = User.query.options(joinedload(User.company)).order_by(User.id.desc()).all()
    companies = Company.query.order_by(Company.name.asc()).all()
    return render_template('admin/manage_users.html', users=users, companies=companies)


# ============================================================
# 🔹 Save New User
# ============================================================
@admin_bp.route('/save-user', methods=['POST'])
@login_required
def save_user():
    """Save new user to database with validation"""

    if getattr(current_user, "role", None) != "admin":
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for('auth.login'))

    # Get form data
    first_name = request.form.get('firstName', '').strip()
    last_name = request.form.get('lastName', '').strip()
    username = request.form.get('username', '').strip().lower()
    email = request.form.get('email', '').strip().lower()
    role = request.form.get('role', '').strip()
    company_id_raw = request.form.get('company', '').strip()
    shift = request.form.get('shift', '').strip()
    salary = request.form.get('salary', '').strip()
    allowance = request.form.get('allowance', 'no').strip().lower()
    allowance_amount = request.form.get('allowance_amount', '').strip()
    default_password = request.form.get('password', 'Default@1234').strip()

    # --- Validation checks ---
    errors = []
    if not first_name: errors.append("First name is required.")
    if not last_name: errors.append("Last name is required.")
    if not username: errors.append("Username is required.")
    if not email: errors.append("Email is required.")
    if not role: errors.append("Role is required.")
    if not company_id_raw: errors.append("Company is required.")
    if not shift: errors.append("Shift is required.")
    if not salary:
        errors.append("Salary is required.")
        salary_val = 0.0
    else:
        try:
            salary_val = float(salary)
            if salary_val < 0:
                errors.append("Salary must be positive.")
        except ValueError:
            errors.append("Salary must be a number.")
            salary_val = 0.0

    # basic email format
    if email and ("@" not in email or "." not in email):
        errors.append("Enter a valid email address.")
    if username and User.query.filter_by(username=username).first():
        errors.append("Username is already taken.")
    if email and User.query.filter_by(email=email).first():
        errors.append("Email is already registered.")

    # travel allowance
    travel_allowance_eligible = (allowance == 'yes')
    travel_allowance_amount_val = None
    if travel_allowance_eligible:
        if not allowance_amount:
            errors.append("Allowance amount required when eligible.")
        else:
            try:
                travel_allowance_amount_val = int(allowance_amount)
                if travel_allowance_amount_val < 0:
                    errors.append("Allowance amount must be non-negative.")
            except ValueError:
                errors.append("Allowance amount must be a number.")

    # Convert company id to int and validate
    company_obj = None
    try:
        company_id = int(company_id_raw)
        company_obj = Company.query.get(company_id)
        if not company_obj:
            errors.append("Invalid company selected.")
    except (ValueError, TypeError):
        errors.append("Invalid company selected.")

    if errors:
        for e in errors:
            flash(e, "danger")
        return redirect(url_for('admin.manage_users'))

    # ✅ Create and save user
    user = User(
        first_name=first_name,
        last_name=last_name,
        username=username,
        email=email,
        role=role.lower(),
        company_id=company_id,  # integer id
        shift=shift,
        salary=salary_val,
        travel_allowance_eligible=travel_allowance_eligible,
        travel_allowance_amount=travel_allowance_amount_val,
        is_active_db=True
    )
    user.set_password(default_password)

    db.session.add(user)
    db.session.commit()

    flash(f"User '{username}' created successfully under company '{company_obj.name}'.", "success")
    return redirect(url_for('admin.manage_users'))


# ============================================================
# 🔹 Get Users for a Specific Company
# ============================================================
@admin_bp.route('/get-company-users/<int:company_id>', methods=['GET'])
@login_required
def get_company_users(company_id):
    """Return list of users enrolled under a company (for modal view)"""
    if getattr(current_user, "role", None) != "admin":
        return {"error": "Unauthorized"}, 403

    company = Company.query.get_or_404(company_id)
    users = User.query.filter_by(company_id=company_id).all()

    data = [
        {
            "id": u.id,
            "name": f"{u.first_name} {u.last_name}",
            "email": u.email,
            "shift": u.shift
        }
        for u in users
    ]
    return {"company": company.name, "users": data}, 200


# --------------------------------------------
# TOGGLE USER ACTIVE/INACTIVE STATUS
# --------------------------------------------
@admin_bp.route('/toggle-user/<int:user_id>', methods=['POST'])
@login_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    # toggle using underlying db column
    user.is_active_db = not user.is_active_db
    db.session.commit()
    flash(f"User '{user.username}' status updated.", "success")
    return redirect(url_for('admin.manage_users'))

@admin_bp.route('/admin/reset-password/<int:user_id>', methods=['POST'])
@login_required
def reset_password(user_id):
    user = User.query.get_or_404(user_id)

    default_password = "Default@1234"
    user.password_hash = generate_password_hash(default_password)

    db.session.commit()

    flash(f"Password for {user.first_name} {user.last_name} has been reset to Default@1234.", "success")
    return redirect(url_for('admin.manage_users'))


# ============================================================
# 🔹 Manage Companies
# ============================================================
@admin_bp.route('/manage-companies')
@login_required
def manage_companies():
    if getattr(current_user, "role", None) != "admin":
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for('auth.login'))

    companies = Company.query.order_by(Company.name.asc()).all()
    return render_template('admin/manage_companies.html', companies=companies)

# ============================================================
# 🔹 Save / Add Company
# ============================================================
@admin_bp.route('/save-company', methods=['POST'])
@login_required
def save_company():
    """Add a new company to the system"""
    if getattr(current_user, "role", None) != "admin":
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for('auth.login'))

    company_name = request.form.get("company_name", "").strip()
    created_by = current_user.username

    if not company_name:
        flash("Company name is required.", "danger")
        return redirect(url_for("admin.manage_companies"))

    existing = Company.query.filter_by(name=company_name).first()
    if existing:
        flash("Company already exists.", "warning")
        return redirect(url_for("admin.manage_companies"))

    new_company = Company(name=company_name, created_by=created_by)
    db.session.add(new_company)
    db.session.commit()

    flash(f"Company '{company_name}' created successfully.", "success")
    return redirect(url_for("admin.manage_companies"))



# ============================================================
# 🔹 Delete Company
# ============================================================
@admin_bp.route('/delete-company', methods=['POST'])
@login_required
def delete_company():
    """Delete company with password check"""
    if getattr(current_user, "role", None) != "admin":
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for('auth.login'))

    company_id = request.form.get('company_id')
    admin_password = request.form.get('admin_password', '')

    # Verify admin password
    if not current_user.check_password(admin_password):
        flash("Incorrect password. Company not deleted.", "danger")
        return redirect(url_for('admin.manage_companies'))

    company = Company.query.get_or_404(company_id)

    # ✅ Count users by company_id, not name
    user_count = User.query.filter_by(company_id=company.id).count()
    if user_count > 0:
        flash(f"Cannot delete '{company.name}' — {user_count} users still assigned.", "warning")
        return redirect(url_for('admin.manage_companies'))

    db.session.delete(company)
    db.session.commit()

    flash(f"Company '{company.name}' deleted successfully.", "success")
    return redirect(url_for('admin.manage_companies'))


@admin_bp.route('/salary-management')
@login_required
def salary_management():
    """Show salary management table with company and shift filters."""
    users = User.query.all()
    companies = Company.query.order_by(Company.name).all()  # ✅ Fetch all companies

    return render_template(
        'admin/salary_management.html',
        users=users,
        companies=companies  # ✅ Pass to template
    )


# 🔹 Add Increment
@admin_bp.route('/add-increment/<int:user_id>', methods=['POST'])
@login_required
def add_increment(user_id):
    user = User.query.get_or_404(user_id)
    increment_amount = float(request.form['increment_amount'])
    reason = request.form['reason']

    previous_salary = user.salary or 0.0
    new_salary = previous_salary + increment_amount

    new_increment = Increment(
        user_id=user.id,
        previous_salary=previous_salary,  # ✅ save old salary
        increment_amount=increment_amount,
        new_salary=new_salary,
        reason=reason
    )
    db.session.add(new_increment)

    # Update user's current salary
    user.salary = new_salary
    db.session.commit()

    flash(f"Salary increment of {increment_amount} for {user.first_name}'s added successfully!", "success")
    return redirect(url_for('admin.salary_management'))


@admin_bp.route('/revoke_increment/<int:increment_id>', methods=['POST'])
def revoke_increment(increment_id):
    increment = Increment.query.get_or_404(increment_id)
    user = increment.user

    try:
        # Begin transaction
        user.salary = increment.previous_salary
        db.session.delete(increment)
        db.session.commit()  # ✅ Commit all at once

        flash(f"Increment revoked successfully. {user.first_name}'s salary reverted to Rs. {user.salary:.0f}.", "success")

    except Exception as e:
        db.session.rollback()  # ❌ Undo all changes if any error occurs
        flash(f"Error revoking increment: {str(e)}", "danger")

    return redirect(url_for('admin.view_increment_history', user_id=user.id))


# 🔹 View Increment History
@admin_bp.route('/increment-history/<int:user_id>')
@login_required
def view_increment_history(user_id):
    user = User.query.get_or_404(user_id)
    increments = Increment.query.filter_by(user_id=user_id).order_by(Increment.date_added.desc()).all()
    return render_template('admin/increment_history.html', user=user, increments=increments)




# -------------------------------
# AJAX Validation Route (Real-time checks)
# -------------------------------
@admin_bp.route('/validate-user-input', methods=['POST'])
@login_required
def validate_user_input():
    """Real-time AJAX validation for username, email, company name, etc."""
    from app.models import Company

    if getattr(current_user, "role", None) != "admin":
        return {"valid": False, "error": "Unauthorized"}, 403

    field = request.json.get("field")
    value = request.json.get("value", "").strip().lower()

    if not field or not value:
        return {"valid": False, "error": "Missing field or value"}, 400

    # --- Username check
    if field == "username":
        exists = User.query.filter(User.username.ilike(value)).first()
        if exists:
            return {"valid": False, "error": "Username already exists."}
        return {"valid": True}

    # --- Email check
    elif field == "email":
        exists = User.query.filter(User.email.ilike(value)).first()
        if exists:
            return {"valid": False, "error": "Email already registered."}
        if "@" not in value or "." not in value:
            return {"valid": False, "error": "Invalid email format."}
        return {"valid": True}

    # --- Company check (optional)
    elif field == "company":
        exists = Company.query.filter(Company.name.ilike(value)).first()
        if not exists:
            return {"valid": False, "error": "Company not found in system."}
        return {"valid": True}

    return {"valid": False, "error": "Unknown field."}, 400

# =====================================
# 🔹 View Users Page
# =====================================
@admin_bp.route('/view-users')
@login_required
def view_users():
    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for('auth.login'))

    users = User.query.options(joinedload(User.company)).order_by(User.id.desc()).all()
    current_time = datetime.now()
    companies = Company.query.order_by(Company.name.asc()).all()
    return render_template('admin/view_users.html', users=users, companies=companies, current_time=current_time)



# =====================================
# 🔹 View Individual User (with penalties)
# =====================================
@admin_bp.route('/view-user/<int:user_id>')
@login_required
def view_user_details(user_id):
    from datetime import datetime
    from collections import defaultdict

    user = User.query.get_or_404(user_id)
    current_time = datetime.utcnow()

    # ✅ Fetch all attendance records for this user
    attendance = (
        Attendance.query
        .filter_by(user_id=user_id)
        .order_by(Attendance.date.desc())
        .all()
    )

    # ✅ Fetch all penalties for this user (with latest first)
    penalties = (
        Penalty.query
        .filter_by(user_id=user_id)
        .order_by(Penalty.created_at.desc())
        .all()
    )

    # ✅ Group penalties by date
    penalties_by_date = defaultdict(list)
    for p in penalties:
        p_date = p.created_at.date() if p.created_at else None
        if p_date:
            penalties_by_date[p_date].append({
                "amount": p.amount,
                "reason": p.reason or "No reason provided",
                "added_by": f"{p.marker.first_name} {p.marker.last_name}" if p.marker else "Unknown"
            })

    # ✅ Attach penalties to each attendance record
    for record in attendance:
        record_date = record.date
        record.penalty_details = penalties_by_date.get(record_date, [])

    # ✅ Attendance stats
    total_presents = sum(1 for r in attendance if r.status.lower() == "present")
    total_lates = sum(
        1 for r in attendance
        if str(r.is_late).lower() in ['true', '1', 'yes', 'late']
    )
    total_absents = sum(1 for r in attendance if r.status.lower() == "absent")
    total_offs = sum(1 for r in attendance if r.status.lower() == "off")

    # ✅ Calculate totals correctly
    total_penalties = sum(p.amount for p in penalties)  # ✅ from penalties table
    total_bonuses = sum(r.bonus or 0 for r in attendance)
    total_salary = user.salary or 0
    pending = total_salary - total_penalties + total_bonuses

    stats = {
        "total_presents": total_presents,
        "total_lates": total_lates,
        "total_absents": total_absents,
        "total_offs": total_offs,
        "penalties": total_penalties,
        "bonuses": total_bonuses,
        "salary": total_salary,
        "pending": pending,
    }

    # ✅ Group attendance records by month
    monthly_data = {}
    for record in attendance:
        month_key = record.date.strftime('%Y-%m')
        monthly_data.setdefault(month_key, []).append(record)

    return render_template(
        'admin/view_user_details.html',
        user=user,
        current_time=current_time,
        attendance=attendance,
        stats=stats,
        monthly_data=monthly_data
    )



# ============================================================
# 🔹 Download All Users (CSV Export)
# ============================================================
import csv
from io import StringIO
from flask import Response

@admin_bp.route('/download-all-users')
@login_required
def download_all_users():
    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for('auth.login'))

    users = User.query.options(joinedload(User.company)).order_by(User.id).all()

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Username", "Email", "Role", "Company", "Shift", "Salary"])

    for user in users:
        writer.writerow([
            user.id,
            f"{user.first_name} {user.last_name}",
            user.username,
            user.email,
            user.role,
            user.company.name if user.company else "",
            user.shift or "",
            user.salary or 0
        ])

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=all_users.csv"
    return response


# ============================================================
# 🔹 Download Single User Attendance (CSV Export)
# ============================================================
@admin_bp.route('/download-user/<int:user_id>')
@login_required
def download_user(user_id):
    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for('auth.login'))

    user = User.query.get_or_404(user_id)
    attendance = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.date.desc()).all()

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Status", "Shift", "Late", "Penalty", "Bonus"])

    for record in attendance:
        writer.writerow([
            record.date.strftime("%Y-%m-%d") if record.date else "",
            record.status or "",
            record.shift or "",
            "Yes" if record.is_late else "No",
            record.penalty or 0,
            record.bonus or 0
        ])

    filename = f"{user.username}_attendance.csv"
    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@admin_bp.route('/get_company_users_json/<int:company_id>')
@login_required
def get_company_users_json(company_id):
    company = Company.query.get_or_404(company_id)
    users_list = []
    for user in company.users:  # Assuming Company.users relationship exists
        users_list.append({
            "id": user.id,
            "username": user.username,
            "full_name": f"{user.first_name} {user.last_name}",
            "salary": f"${user.salary}",
            "joined_at": user.created_at.strftime("%Y-%m-%d %H:%M")
        })
    return jsonify({"users": users_list})

# -------------------------------
# ATTENDANCE MANAGEMENT PAGE (Admin)
# -------------------------------
# -------------------------------
# ATTENDANCE MANAGEMENT PAGE (Multi-level Expandable)
# -------------------------------
@admin_bp.route('/attendance', methods=['GET'])
@login_required
def attendance():
    from calendar import month_name
    from datetime import date, datetime, timedelta
    from collections import defaultdict
    import calendar

    # Get filters
    month_filter = request.args.get('month', '')
    shift_filter = request.args.get('shift', '').lower().strip()

    # If no month given, use current month (YYYY-MM)
    if not month_filter:
        month_filter = date.today().strftime("%Y-%m")

    # Convert to actual date range for PostgreSQL
    year, month = map(int, month_filter.split('-'))
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    # Determine visible companies
    if current_user.role.lower() == "admin":
        companies = Company.query.order_by(Company.name.asc()).all()
    else:
        companies = [current_user.company] if current_user.company else []

    attendance_data = {}

    for company in companies:
        # ✅ Efficient date range filter (PostgreSQL optimized)
        query = (
            Attendance.query
            .join(User, Attendance.user_id == User.id)
            .filter(User.company_id == company.id)
            .filter(Attendance.date >= start_date)
            .filter(Attendance.date <= end_date)
        )

        if shift_filter:
            query = query.filter(User.shift.ilike(shift_filter))

        records = query.order_by(Attendance.date.desc(), Attendance.time.desc()).all()
        if not records:
            continue

        # --- Grouping ---
        grouped_by_month = defaultdict(lambda: defaultdict(list))
        for r in records:
            month_key = r.date.strftime("%Y-%m")
            grouped_by_month[month_key][r.date].append(r)

        # --- Build Structure ---
        attendance_data[company.id] = {
            "company": company,
            "months": {}
        }

        for month_key, date_records in grouped_by_month.items():
            month_label = f"{month_name[int(month_key.split('-')[1])]} {month_key.split('-')[0]}"

            month_summary = {
                "label": month_label,
                "dates": [],
                "summary": {
                    "presents": 0,
                    "lates": 0,
                    "absents": 0,
                    "offs": 0,
                    "bonuses": 0.0,
                    "penalties": 0.0,
                },
            }

            for dt, records_for_date in sorted(date_records.items(), reverse=True):
                day_summary = {
                    "date": dt,
                    "records": [],
                    "presents": 0,
                    "lates": 0,
                    "absents": 0,
                    "offs": 0,
                    "bonuses": 0.0,
                    "penalties": 0.0,
                }

                for r in records_for_date:
                    record_info = {
                        "user": r.user,
                        "status": r.status,
                        "is_late": r.is_late,
                        "time": r.time.strftime("%H:%M"),
                        "bonus": r.bonus,
                        "penalty": r.penalty,
                        "marked_by": r.marker.user_full_name() if r.marker else "—",
                        "shift": r.user.shift or "-",
                    }
                    day_summary["records"].append(record_info)

                    # Tally per day
                    if r.status == "Present":
                        day_summary["presents"] += 1
                    elif r.status == "Late":
                        day_summary["lates"] += 1
                    elif r.status == "Absent":
                        day_summary["absents"] += 1
                    elif r.status == "Off":
                        day_summary["offs"] += 1
                    day_summary["bonuses"] += r.bonus or 0
                    day_summary["penalties"] += r.penalty or 0

                # Add to month summary
                month_summary["dates"].append(day_summary)
                month_summary["summary"]["presents"] += day_summary["presents"]
                month_summary["summary"]["lates"] += day_summary["lates"]
                month_summary["summary"]["absents"] += day_summary["absents"]
                month_summary["summary"]["offs"] += day_summary["offs"]
                month_summary["summary"]["bonuses"] += day_summary["bonuses"]
                month_summary["summary"]["penalties"] += day_summary["penalties"]

            attendance_data[company.id]["months"][month_key] = month_summary

    # Fetch recent clearances (optional)
    clearances = Clearance.query.order_by(Clearance.date_added.desc()).limit(10).all()

    # Shifts for dropdown
    shifts = ["morning", "evening", "night"]

    return render_template(
        "admin/attendance.html",
        attendance_data=attendance_data,
        month_filter=month_filter,
        shift_filter=shift_filter,
        shifts=shifts,
        current_date=date.today(),
        clearances=clearances,
    )

# ==========================================================
# 📤 DOWNLOAD ATTENDANCE REPORT (EXCEL) - company + month
# ==========================================================
@admin_bp.route('/download_attendance_report/<int:company_id>/<string:month>', methods=['GET'])
@login_required
def download_attendance_report(company_id, month):
    """
    month: "YYYY-MM" string
    Generates an .xlsx containing:
    Sr#, Full Name, Role, Shift, <date cols for month>, Presents, Lates, Absents, Offs,
    Totals, Base Salary, Salary(from days), Bonus, Penalty, Clearance, Calculated Salary
    """
    from io import BytesIO
    from datetime import date
    from calendar import monthrange
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file, flash, redirect, url_for
    from sqlalchemy import extract

    # permission checks
    role = (current_user.role or "").lower()
    if role not in ("admin", "supervisor"):
        flash("Access denied: only admins or supervisors can download reports.", "danger")
        return redirect(url_for('admin.attendance'))

    company = Company.query.get_or_404(company_id)
    if role == "supervisor" and company.id != current_user.company_id:
        flash("You can only download your own company's report.", "warning")
        return redirect(url_for('admin.attendance'))

    # parse month -> year, month_num
    try:
        year_str, month_str = month.split("-")
        year = int(year_str)
        month_num = int(month_str)
    except Exception:
        flash("Invalid month format. Use YYYY-MM.", "danger")
        return redirect(url_for('admin.attendance'))

    # Build list of all dates in the month
    num_days = monthrange(year, month_num)[1]
    date_list = [date(year, month_num, d) for d in range(1, num_days + 1)]

    # Fetch users in company (agents + supervisors)
    users = (
        User.query
        .filter(User.company_id == company.id)
        .filter(User.role.in_(["agent", "supervisor"]))
        .all()
    )

    user_ids = [u.id for u in users] or [0]

    # Attendance, penalties, clearances
    attendances = (
        Attendance.query
        .filter(Attendance.user_id.in_(user_ids))
        .filter(extract('year', Attendance.date) == year)
        .filter(extract('month', Attendance.date) == month_num)
        .all()
    )

    penalties = (
        Penalty.query
        .filter(Penalty.user_id.in_(user_ids))
        .filter(extract('year', Penalty.created_at) == year)
        .filter(extract('month', Penalty.created_at) == month_num)
        .all()
    )

    clearances = (
        Clearance.query
        .filter(Clearance.user_id.in_(user_ids))
        .filter(extract('year', Clearance.date_added) == year)
        .filter(extract('month', Clearance.date_added) == month_num)
        .all()
    )

    # Index data
    att_by_user_date = defaultdict(list)
    for a in attendances:
        att_by_user_date[(a.user_id, a.date)].append(a)

    penalty_by_user = defaultdict(list)
    for p in penalties:
        penalty_by_user[p.user_id].append(p)

    clearance_by_user = defaultdict(list)
    for c in clearances:
        clearance_by_user[c.user_id].append(c)

    # Group users by shift
    users_by_shift = defaultdict(list)
    for u in users:
        users_by_shift[u.shift or "Unassigned"].append(u)

    # Sort: supervisors first, then agents, with separator rows
    ordered_users = []
    for shift_name in sorted(users_by_shift.keys()):
        shift_users = users_by_shift[shift_name]
        supervisors = sorted([x for x in shift_users if x.role == "supervisor"], key=lambda y: (y.first_name or "", y.last_name or ""))
        agents = sorted([x for x in shift_users if x.role == "agent"], key=lambda y: (y.first_name or "", y.last_name or ""))
        ordered_users.extend([(shift_name, s) for s in supervisors])
        ordered_users.extend([(shift_name, a) for a in agents])
        ordered_users.append((shift_name, None))  # shift separator

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{company.name[:25]}_{month}"

    # Header row
    header = ["Sr No", "Full Name", "Role", "Shift"]
    header += [dt.strftime("%d") for dt in date_list]
    header += [
        "Presents", "Lates", "Absents", "Offs", "Totals",
        "Base Salary", "Salary(from days)",
        "Bonus", "Penalty", "Clearance", "Calculated Salary"
    ]
    ws.append(header)

    # Styling header
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_align

    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    sr = 1
    for shift_name, user in ordered_users:
        if user is None:
            # Separator row
            sep_idx = ws.max_row + 1
            ws.append([""] * len(header))
            for c in range(1, len(header) + 1):
                ws.cell(row=sep_idx, column=c).fill = black_fill
            continue

        row = [sr, user.full_name, user.role.capitalize(), shift_name]
        presents = lates = absents = offs = 0
        total_bonus = total_penalty = 0.0

        # Date cells
        for dt in date_list:
            marks = att_by_user_date.get((user.id, dt), [])
            if not marks:
                row.append("")
                continue

            statuses = [m.status for m in marks]
            if "Present" in statuses:
                row.append("P")
                presents += 1
            elif "Late" in statuses:
                row.append("L")
                lates += 1
            elif "Off" in statuses:
                row.append("Off")
                offs += 1
            elif "Absent" in statuses:
                row.append("A")
                absents += 1
            else:
                row.append(statuses[0])

            for m in marks:
                total_bonus += (m.bonus or 0)
                total_penalty += (m.penalty or 0)

        totals = presents + lates + absents + offs
        base_salary = user.salary or 0
        per_day_salary = base_salary / num_days

        # Updated salary logic
        salary_from_days = (
            (presents * per_day_salary)
            + (lates * (per_day_salary - 400))
            + (offs * per_day_salary)
        )

        # Monthly penalties/clearances
        penalty_extra = sum(p.amount for p in penalty_by_user.get(user.id, []))
        clearance_extra = sum(c.amount for c in clearance_by_user.get(user.id, []))

        total_penalty_combined = total_penalty + penalty_extra
        total_bonus_combined = total_bonus
        total_clearance_combined = clearance_extra

        calculated_salary = (
            salary_from_days
            + total_bonus_combined
            - total_penalty_combined
            + total_clearance_combined
        )

        row += [
            presents, lates, absents, offs, totals,
            round(base_salary, 2),
            round(salary_from_days, 2),
            round(total_bonus_combined, 2),
            round(total_penalty_combined, 2),
            round(total_clearance_combined, 2),
            round(calculated_salary, 2),
        ]

        ws.append(row)
        sr += 1

    # Freeze header row + first 4 columns
    ws.freeze_panes = "E2"

    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max_len + 2

    # Output
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = company.name.replace(" ", "_")
    filename = f"{safe_name}_Attendance_{month}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# -------------------------------
# AJAX FILTER ENDPOINT
# -------------------------------
@admin_bp.route('/attendance/filter')
@login_required
def attendance_filter():
    company_id = request.args.get('company_id', type=int)
    month = request.args.get('month')
    shift = request.args.get('shift')

    records = (
        Attendance.query
        .join(User)
        .filter(User.company_id == company_id)
    )

    if month:
        records = records.filter(Attendance.date.like(f"{month}-%"))
    if shift:
        records = records.filter(User.shift == shift)

    records = records.order_by(Attendance.date.desc()).all()

    return render_template('admin/partials/attendance_table.html', records=records)


@admin_bp.route('/unlock_profile/<int:user_id>', methods=['POST'])
@login_required
def unlock_profile(user_id):
    if current_user.role != 'admin':
        flash("Unauthorized action.", "danger")
        return redirect(url_for('admin.manage_users'))

    user = User.query.get_or_404(user_id)
    
    # Toggle lock/unlock
    user.profile_locked = not user.profile_locked
    db.session.commit()

    status = "unlocked" if not user.profile_locked else "locked"
    flash(f"Profile for {user.username} has been {status} successfully!", "success")
    return redirect(url_for('admin.manage_users'))



# ==========================================================
# ADD PENALTY (Admin & Supervisor)
# ==========================================================
@admin_bp.route("/penalties/add", methods=["GET", "POST"])
@login_required
def add_penalty():
    # Only admin or supervisor can access
    if current_user.role not in ["admin", "supervisor"]:
        flash("Access denied.", "danger")
        return redirect(url_for("auth.login"))

    # Admin can see all agents & supervisors
    query = User.query.filter(User.role.in_(["agent", "supervisor"]))

    if current_user.role == "supervisor":
        # Supervisors can only see agents of their own company
        query = query.filter(User.company_id == current_user.company_id, User.role=="agent")

    # Optional search filters
    search_name = request.args.get("search_name")
    search_username = request.args.get("search_username")
    if search_name:
        query = query.filter(User.full_name.ilike(f"%{search_name}%"))
    if search_username:
        query = query.filter(User.username.ilike(f"%{search_username}%"))

    users = query.order_by(User.full_name.asc()).all()

    if request.method == "POST":
        user_id = request.form.get("user_id")
        amount = request.form.get("amount")
        reason = request.form.get("reason")

        if not user_id or not amount or not reason:
            flash("All fields are required.", "danger")
            return redirect(url_for("supervisor.add_penalty"))

        penalty = Penalty(
            user_id=int(user_id),
            amount=float(amount),
            reason=reason,
            added_by_id=current_user.id
        )
        db.session.add(penalty)
        db.session.commit()
        flash("Penalty added successfully.", "success")
        return redirect(request.url)

    return render_template(
        "supervisor/add_penalty.html",
        users=users,
        current_user=current_user
    )
